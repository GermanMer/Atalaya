#PRIMER PASO (Obtener los datos)
dia = input('Ingrese el día:')
mes = input('Ingrese el mes:')
año = input('Ingrese el año:')

totales = list()
fhand = open('p.txt')

for linea in fhand:
#Totales surtidores
    if linea.startswith('                          00-00002  DIESEL X10                                  '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[3])
    if linea.startswith('                          00-00003  QUANTIUM DIESEL                             '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[3])
    if linea.startswith('                          00-00004  GNC                                         '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[2])
    if linea.startswith('                          00-00005  QUANTIUM PREMIUM                            '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[3])
    if linea.startswith('                          00-00006  ENERGY 5000                                 '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[3])
    if linea.startswith('                          00-01232  DIESEL GRADO 2                              '):
        linea = linea.rstrip()
        palabras = linea.split()
        try: totales.append(palabras[4])
        except: continue
#print(totales)

nuevototales = [v.replace(',', '.') for v in totales]

#dieseltotal = totales[0]

#try:
#    dieseltotal = float(nuevototales[0]) + float(nuevototales[5])
    #print(dieseltotal)
#except: pass

#SEGUNDO PASO (completar los datos)
import os
os.chdir('Z:\excel\PLANILLAS DE FACTURACION DE COMBUSTIBLES\\'+año)

import openpyxl
wb = openpyxl.load_workbook(mes+' FACTURACION DE COMBUSTIBLES.xlsx')
hoja = wb["Existencias y Ventas Mensuales"]

qp = hoja.cell(row=int(dia)+5, column=3, value=float(nuevototales[3]))
ns = hoja.cell(row=int(dia)+5, column=10, value=float(nuevototales[4]))
qd = hoja.cell(row=int(dia)+5, column=17, value=float(nuevototales[1]))
ds = hoja.cell(row=int(dia)+5, column=24, value=float(nuevototales[0]))
am = hoja.cell(row=int(dia)+5, column=31, value=float(nuevototales[2]))

wb.save(mes+' FACTURACION DE COMBUSTIBLES.xlsx')
