#PRIMER PASO (Recuperar los datos del archivo CAJA DIARIA del día)

dia = input('Ingrese el día:')
mes = input('Ingrese el mes:')
año = input('Ingrese el año:')

totales = list()
cuentacorriente = list()
fhand = open('p.txt')

for linea in fhand:

#Totales surtidores
    if linea.startswith('                          00-00002  DIESEL X10                                  '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[4])
    if linea.startswith('                          00-00003  QUANTIUM DIESEL                             '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[4])
    if linea.startswith('                          00-00004  GNC                                         '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[3])
    if linea.startswith('                          00-00005  QUANTIUM PREMIUM                            '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[4])
    if linea.startswith('                          00-00006  ENERGY 5000                                 '):
        linea = linea.rstrip()
        palabras = linea.split()
        totales.append(palabras[4])
    if linea.startswith('                          00-01232  DIESEL GRADO 2                              '):
        linea = linea.rstrip()
        palabras = linea.split()
        try: totales.append(palabras[5])
        except: continue

#Cuenta Corriente surtidores
    try:
        if linea.startswith('                 00-00002 DIESEL X10'):
            linea = linea.rstrip()
            palabras = linea.split()
            cuentacorriente.append(palabras[6])
    except: cuentacorriente.append("0,00")
    try:
        if linea.startswith('                 00-00003 QUANTIUM DIESEL'):
            linea = linea.rstrip()
            palabras = linea.split()
            cuentacorriente.append(palabras[6])
    except: cuentacorriente.append("0,00")
    if linea.startswith('                 00-00004 GNC'):
        linea = linea.rstrip()
        palabras = linea.split()
        cuentacorriente.append(palabras[5])
    if linea.startswith('                 00-00005 QUANTIUM PREMIUM'):
        linea = linea.rstrip()
        palabras = linea.split()
        cuentacorriente.append(palabras[6])
    if linea.startswith('                 00-00006 ENERGY 5000'):
        linea = linea.rstrip()
        palabras = linea.split()
        cuentacorriente.append(palabras[6])
    #if linea.startswith('                 00-01232 GASOIL ADITIVADO GR'):
    #    linea = linea.rstrip()
    #    palabras = linea.split()
    #    try: cuentacorriente.append(palabras[5])
    #    except: continue

#print(totales)
nuevototales = [v.replace(',', '.') for v in totales]
#print(nuevototales)

#print(cuentacorriente)
nuevocuentacorriente = [v2.replace(',', '.') for v2 in cuentacorriente]
#print(nuevocuentacorriente)

dieselcontado = float(nuevototales[0]) - float(nuevocuentacorriente[0])
ndieselcontado = round(dieselcontado, 2)
#print(ndieselcontado)

quantiumdcontado = float(nuevototales[1]) - float(nuevocuentacorriente[1])
nquantiumdcontado = round(quantiumdcontado, 2)
#print(nquantiumdcontado)

gnccontado = float(nuevototales[2]) - float(nuevocuentacorriente[2])
ngnccontado = round(gnccontado, 2)
#print(ngnccontado)

quantiumpcontado = float(nuevototales[3]) - float(nuevocuentacorriente[3])
nquantiumpcontado = round(quantiumpcontado, 2)
#print(nquantiumpcontado)

naftasuper = float(nuevototales[4]) - float(nuevocuentacorriente[4])
nnaftasuper = round(naftasuper, 2)
#print(nnaftasuper)

#try:
#    dieseltr = float(nuevototales[5]) - float(nuevocuentacorriente[5])
#    ndieseltr = round(dieseltr, 2)
#    #print(ndieseltr)
#except: pass

#dieselsumados
#try:
#    dieselsumadoscontado = ndieselcontado + ndieseltr
    #print(dieselsumadoscontado)
#except: pass

#try:
#    dieselsumadoscc = float(nuevocuentacorriente[0]) + float(nuevocuentacorriente[5])
    #print(dieselsumadoscc)
#except: pass


#SEGUNDO PASO (Tomar mas datos del Excel CAJA DIARIA)
import os
os.chdir('Z:\excel\A. CAJA DIARIA\\'+año+'\\'+mes)

import openpyxl
wb = openpyxl.load_workbook('CAJA DIARIA '+dia+'-'+mes+'.xlsx')
hoja2 = wb.active

#Lubricantes
lubricantescc = hoja2['P34'].value
#print(lubricantescc)
b58 = hoja2['B58'].value
b59 = hoja2['B59'].value
b60 = hoja2['B60'].value
lubricantesct = b58 + b59 + b60 - lubricantescc
#print(lubricantesct)

#Accesorios
accesorioscc = hoja2['P35'].value
#print(accesorioscc)
c58 = hoja2['C58'].value
c59 = hoja2['C59'].value
c60 = hoja2['C60'].value
accesoriosct = c58 + c59 + c60 - accesorioscc
#print(accesoriosct)

#Servicios
servicioscc = hoja2['P36'].value
#print(servicioscc)
e58 = hoja2['E58'].value
e59 = hoja2['E59'].value
e60 = hoja2['E60'].value
serviciosct = e58 + e59 + e60 - servicioscc
#print(serviciosct)

#Axionmix
precioamionmix = hoja2['O24'].value
axionmixcc = hoja2['P37'].value
k43 = hoja2['K43'].value
k44 = hoja2['K44'].value
axionmixct = ((k43 - k44) * precioamionmix) - axionmixcc
#print(axionmixct)


#TERCER PASO (Cargar los datos en RESUMEN CDO Y CTA CTE)
wb = openpyxl.load_workbook('RESUMEN CDO Y CTA CTE.xlsx')
hoja = wb["Hoja1"]

x10cc = hoja.cell(row=int(dia)+5, column=2, value=float(nuevocuentacorriente[0]))
#try:
#    x10cc = hoja.cell(row=int(dia)+5, column=2, value=dieselsumadoscc)
#except: pass
x10c = hoja.cell(row=int(dia)+5, column=3, value=ndieselcontado)
#try:
#    x10c = hoja.cell(row=int(dia)+5, column=3, value=dieselsumadoscontado)
#except: pass
qdcc = hoja.cell(row=int(dia)+5, column=4, value=float(nuevocuentacorriente[1]))
qdc = hoja.cell(row=int(dia)+5, column=5, value=nquantiumdcontado)
gnccc = hoja.cell(row=int(dia)+5, column=6, value=float(nuevocuentacorriente[2]))
gncc = hoja.cell(row=int(dia)+5, column=7, value=ngnccontado)
qpcc = hoja.cell(row=int(dia)+5, column=8, value=float(nuevocuentacorriente[3]))
pqc = hoja.cell(row=int(dia)+5, column=9, value=nquantiumpcontado)
nscc = hoja.cell(row=int(dia)+5, column=10, value=float(nuevocuentacorriente[4]))
nsc = hoja.cell(row=int(dia)+5, column=11, value=nnaftasuper)

lcc = hoja.cell(row=int(dia)+5, column=12, value=lubricantescc)
lct = hoja.cell(row=int(dia)+5, column=13, value=lubricantesct)
accc = hoja.cell(row=int(dia)+5, column=14, value=accesorioscc)
acct = hoja.cell(row=int(dia)+5, column=15, value=accesoriosct)
scc = hoja.cell(row=int(dia)+5, column=16, value=servicioscc)
sct = hoja.cell(row=int(dia)+5, column=17, value=serviciosct)
amcc = hoja.cell(row=int(dia)+5, column=18, value=axionmixcc)
amct = hoja.cell(row=int(dia)+5, column=19, value=axionmixct)

wb.save('RESUMEN CDO Y CTA CTE.xlsx')
