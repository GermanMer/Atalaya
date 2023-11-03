#PRIMER PASO (Recuperar los datos del archivo CAJA DIARIA del día)

import pandas as pd
import openpyxl
import os

dia = input('Ingrese el día:')
mes = input('Ingrese el mes:')
año = input('Ingrese el año:')

os.chdir('Z:\excel\A. CAJA DIARIA\\'+año+'\\'+mes)

nombre_caja = "CAJA DIARIA "+str(dia)+"-"+str(mes)+".xlsx"

caja = pd.read_excel(nombre_caja)

caja.loc[28, 'Unnamed: 15']=caja.loc[28, 'Unnamed: 15'].replace(',', '.')
caja.loc[29, 'Unnamed: 15']=caja.loc[29, 'Unnamed: 15'].replace(',', '.')
caja.loc[30, 'Unnamed: 15']=caja.loc[30, 'Unnamed: 15'].replace(',', '.')
caja.loc[31, 'Unnamed: 15']=caja.loc[31, 'Unnamed: 15'].replace(',', '.')


caja['Unnamed: 15'] = pd.to_numeric(caja['Unnamed: 15'], errors='coerce')
caja['Unnamed: 16'] = pd.to_numeric(caja['Unnamed: 16'], errors='coerce')

x10_cc = caja.loc[27, 'Unnamed: 15']
x10 = caja.loc[27, 'Unnamed: 16']
qd_cc = caja.loc[28, 'Unnamed: 15']
qd = caja.loc[28, 'Unnamed: 16']
gnc_cc = caja.loc[29, 'Unnamed: 15']
gnc = caja.loc[29, 'Unnamed: 16']
qn_cc = caja.loc[30, 'Unnamed: 15']
qn = caja.loc[30, 'Unnamed: 16']
ns_cc = caja.loc[31, 'Unnamed: 15']
ns = caja.loc[31, 'Unnamed: 16']
lub_cc = caja.loc[32, 'Unnamed: 15']
lub = caja.loc[32, 'Unnamed: 16']
acc_cc = caja.loc[33, 'Unnamed: 15']
acc = caja.loc[33, 'Unnamed: 16']
serv_cc = caja.loc[34, 'Unnamed: 15']
serv = caja.loc[34, 'Unnamed: 16']
amix_cc = caja.loc[35, 'Unnamed: 15']
amix = caja.loc[35, 'Unnamed: 16']

#SEGUNDO PASO (Cargar los datos en RESUMEN CDO Y CTA CTE)
wb = openpyxl.load_workbook('RESUMEN CDO Y CTA CTE.xlsx')
hoja = wb["Hoja1"]

hoja.cell(row=int(dia)+5, column=2, value=x10_cc)
hoja.cell(row=int(dia)+5, column=3, value=x10)
hoja.cell(row=int(dia)+5, column=4, value=qd_cc)
hoja.cell(row=int(dia)+5, column=5, value=qd)
hoja.cell(row=int(dia)+5, column=6, value=gnc_cc)
hoja.cell(row=int(dia)+5, column=7, value=gnc)
hoja.cell(row=int(dia)+5, column=8, value=qn_cc)
hoja.cell(row=int(dia)+5, column=9, value=qn)
hoja.cell(row=int(dia)+5, column=10, value=ns_cc)
hoja.cell(row=int(dia)+5, column=11, value=ns)
hoja.cell(row=int(dia)+5, column=12, value=lub_cc)
hoja.cell(row=int(dia)+5, column=13, value=lub)
hoja.cell(row=int(dia)+5, column=14, value=acc_cc)
hoja.cell(row=int(dia)+5, column=15, value=acc)
hoja.cell(row=int(dia)+5, column=16, value=serv_cc)
hoja.cell(row=int(dia)+5, column=17, value=serv)
hoja.cell(row=int(dia)+5, column=18, value=amix_cc)
hoja.cell(row=int(dia)+5, column=19, value=amix)

wb.save('RESUMEN CDO Y CTA CTE.xlsx')
