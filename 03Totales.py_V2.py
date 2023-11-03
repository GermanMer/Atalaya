#PRIMER PASO (Recuperar los datos del archivo CAJA DIARIA del día)

import pandas as pd
import openpyxl
import os

dia = input('Ingrese el día:')
mes = input('Ingrese el mes:')
año = input('Ingrese el año:')

os.chdir('Z:\excel\A. CAJA DIARIA\\'+año+'\\'+mes)

nombre_caja = "CAJA DIARIA "+str(dia)+"-"+str(mes)+".xlsx"

caja = pd.read_excel(nombre_caja)

tqp = caja.loc[17, 'Unnamed: 15']
tns = caja.loc[18, 'Unnamed: 15']
tqd = caja.loc[15, 'Unnamed: 15']
tds = caja.loc[14, 'Unnamed: 15']
tgn = caja.loc[16, 'Unnamed: 15']


#SEGUNDO PASO (completar los datos)

os.chdir('Z:\excel\PLANILLAS DE FACTURACION DE COMBUSTIBLES\\'+año)

wb = openpyxl.load_workbook(mes+' FACTURACION DE COMBUSTIBLES.xlsx')
hoja = wb["Existencias y Ventas Mensuales"]

hoja.cell(row=int(dia)+5, column=3, value=tqp)
hoja.cell(row=int(dia)+5, column=10, value=tns)
hoja.cell(row=int(dia)+5, column=17, value=tqd)
hoja.cell(row=int(dia)+5, column=24, value=tds)
hoja.cell(row=int(dia)+5, column=31, value=tgn)

wb.save(mes+' FACTURACION DE COMBUSTIBLES.xlsx')
