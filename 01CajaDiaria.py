#PRIMER PASO (Recuperar los medidores de la planilla DEBO)

medidoresiniciales = list()
gncydieseltiniciales = list()
fecha = dict()
fhand = open('m.txt')
for linea in fhand:
    if linea.startswith('         $'):
        linea = linea.rstrip()
        palabras = linea.split()
        #print(palabras[4])
        medidoresiniciales.append(palabras[4])
        gncydieseltiniciales.append(palabras[3])
#print('Medidores Iniciales:', medidoresiniciales) #OJO GNC y Diesel T!
#print('Medidores GNC y Diesel T Iniciales:', gncydieseltiniciales)
#FECHA
    if linea.startswith('           RUTA NAC N 9 KM 463'):
        linea = linea.rstrip()
        palabrasf = linea.split()
        fecha['dia'] = palabrasf[9]
        fecha['mes'] = palabrasf[10]
        fecha['año'] = palabrasf[11]

dia = fecha['dia']

if fecha['mes'] == 'enero':
    fecha['mes'] = '01'
if fecha['mes'] == 'febrero':
    fecha['mes'] = '02'
if fecha['mes'] == 'marzo':
    fecha['mes'] = '03'
if fecha['mes'] == 'abril':
    fecha['mes'] = '04'
if fecha['mes'] == 'mayo':
    fecha['mes'] = '05'
if fecha['mes'] == 'junio':
    fecha['mes'] = '06'
if fecha['mes'] == 'julio':
    fecha['mes'] = '07'
if fecha['mes'] == 'agosto':
    fecha['mes'] = '08'
if fecha['mes'] == 'septiembre':
    fecha['mes'] = '09'
if fecha['mes'] == 'octubre':
    fecha['mes'] = '10'
if fecha['mes'] == 'noviembre':
    fecha['mes'] = '11'
if fecha['mes'] == 'diciembre':
    fecha['mes'] = '12'

mes = fecha['mes']
año = fecha['año']
#print(dia+'/'+mes+'/'+año)

medidoresfinales = list()
gncydieseltfinales = list()
fhand2 = open('n.txt')
for linea2 in fhand2:
    if linea2.startswith('         $'):
        linea2 = linea2.rstrip()
        palabras2 = linea2.split()
        #print(palabras2[5])
        medidoresfinales.append(palabras2[5])
        gncydieseltfinales.append(palabras2[4])
#print('Medidores Finales:', medidoresfinales) #OJO GNC y Diesel T!
#print('Medidores GNC y Diesel T Finales:', gncydieseltfinales)

fhand3 = open('p.txt')

axionmix = 0
dieselgrado2 = 0


#Otros despachos:
oddieselx10=0
odquantiumd=0
odquantiump=0
odenergy5000=0

for linea3 in fhand3:
    try:
        if linea3.startswith('                       00-00002   DIESEL X10'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            oddieselx10 = (palabras3[3])
            oddieselx10 = oddieselx10.replace(',', '.')
            oddieselx10 = float(oddieselx10)
    except: continue
    try:
        if linea3.startswith('                       00-00003   QUANTIUM DIESEL'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            odquantiumd = palabras3[3]
            odquantiumd = odquantiumd.replace(',', '.')
            odquantiumd = float(odquantiumd)
    except: continue
#print(odquantiumd)
    try:
        if linea3.startswith('                       00-00005   QUANTIUM PREMIUM'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            odquantiump = palabras3[3]
            odquantiump = odquantiump.replace(',', '.')
            odquantiump = float(odquantiump)
    except: continue
#print(odquantiump)
    try:
        if linea3.startswith('                       00-00006   ENERGY 5000'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            odenergy5000 = palabras3[3]
            odenergy5000 = odenergy5000.replace(',', '.')
            odenergy5000 = float(odenergy5000)
    except: continue
#print(odenergy5000)

#Ventas de combustibles en Cuenta Corriente
    try:
        if linea3.startswith('                 00-00001 AXION MIX'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            axionmix = palabras3[6]
            axionmix = axionmix.replace(',', '.')
            axionmix = float(axionmix)
    except: continue
#print(axionmix)
    try:
        if linea3.startswith('                 00-00002 DIESEL X10'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            dieselx10 = (palabras3[6])
    except: continue
#print(dieselx10)
    try:
        if linea3.startswith('                 00-00003 QUANTIUM DIESEL'):
            linea3 = linea3.rstrip()
            palabras3 = linea3.split()
            quantiumd = palabras3[6]
    except: continue
#print(quantiumd)
    if linea3.startswith('                 00-00004 GNC'):
        linea3 = linea3.rstrip()
        palabras3 = linea3.split()
        gnc = palabras3[5]
#print(gnc)
    if linea3.startswith('                 00-00005 QUANTIUM PREMIUM'):
        linea3 = linea3.rstrip()
        palabras3 = linea3.split()
        quantiump = palabras3[6]
#print(quantiump)
    if linea3.startswith('                 00-00006 ENERGY 5000'):
        linea3 = linea3.rstrip()
        palabras3 = linea3.split()
        energy5000 = palabras3[6]
#print(energy5000)

#    try:
#        if linea3.startswith('                 00-01232 GASOIL ADITIVADO GR'):
#            linea3 = linea3.rstrip()
#            palabras3 = linea3.split()
#            dieselgrado2 = palabras3[5]
#    except: continue
#print(dieselgrado2)

#print(dieselx10, quantiumd, gnc, quantiump, energy5000, axionmix, dieselgrado2)

#SEGUNDO PASO (Completar los valores en el archivo de Excel)
#pip install openpyxl (Ejecutar en el modo interactivo la 1° vez!)

import openpyxl
wb = openpyxl.load_workbook('CAJADIARIA.xlsx')
hoja = wb.active

#Medidores iniciales
b4 = hoja.cell(row=4, column=2, value=medidoresiniciales[0])
b8 = hoja.cell(row=8, column=2, value=medidoresiniciales[1])
b12 = hoja.cell(row=12, column=2, value=medidoresiniciales[2])
b16 = hoja.cell(row=16, column=2, value=medidoresiniciales[3])
b20 = hoja.cell(row=20, column=2, value=medidoresiniciales[4])
b24 = hoja.cell(row=24, column=2, value=medidoresiniciales[5])
b28 = hoja.cell(row=28, column=2, value=medidoresiniciales[6])
b32 = hoja.cell(row=32, column=2, value=medidoresiniciales[7])
b36 = hoja.cell(row=36, column=2, value=medidoresiniciales[8])
b40 = hoja.cell(row=40, column=2, value=medidoresiniciales[9])
b44 = hoja.cell(row=44, column=2, value=medidoresiniciales[10])
b48 = hoja.cell(row=48, column=2, value=medidoresiniciales[11])
#b44 = hoja.cell(row=44, column=2, value=gncydieseltiniciales[44])
#b48 = hoja.cell(row=48, column=2, value=gncydieseltiniciales[45])

e4 = hoja.cell(row=4, column=5, value=medidoresiniciales[12])
e8 = hoja.cell(row=8, column=5, value=medidoresiniciales[13])
e12 = hoja.cell(row=12, column=5, value=medidoresiniciales[14])
e16 = hoja.cell(row=16, column=5, value=medidoresiniciales[15])
e20 = hoja.cell(row=20, column=5, value=medidoresiniciales[16])
e24 = hoja.cell(row=24, column=5, value=medidoresiniciales[17])
e28 = hoja.cell(row=28, column=5, value=medidoresiniciales[18])
e32 = hoja.cell(row=32, column=5, value=medidoresiniciales[19])
e36 = hoja.cell(row=36, column=5, value=medidoresiniciales[20])
e40 = hoja.cell(row=40, column=5, value=medidoresiniciales[21])
e44 = hoja.cell(row=44, column=5, value=medidoresiniciales[22])
e48 = hoja.cell(row=48, column=5, value=medidoresiniciales[23])

h4 = hoja.cell(row=4, column=8, value=gncydieseltiniciales[24])
h8 = hoja.cell(row=8, column=8, value=gncydieseltiniciales[25])
h12 = hoja.cell(row=12, column=8, value=gncydieseltiniciales[26])
h16 = hoja.cell(row=16, column=8, value=gncydieseltiniciales[27])

h24 = hoja.cell(row=24, column=8, value=medidoresiniciales[28])
h28 = hoja.cell(row=28, column=8, value=medidoresiniciales[29])
h32 = hoja.cell(row=32, column=8, value=medidoresiniciales[30])
h36 = hoja.cell(row=36, column=8, value=medidoresiniciales[31])
h40 = hoja.cell(row=40, column=8, value=medidoresiniciales[32])
h44 = hoja.cell(row=44, column=8, value=medidoresiniciales[33])
h48 = hoja.cell(row=48, column=8, value=medidoresiniciales[34])
h52 = hoja.cell(row=52, column=8, value=medidoresiniciales[35])

k4 = hoja.cell(row=4, column=11, value=medidoresiniciales[36])
k8 = hoja.cell(row=8, column=11, value=medidoresiniciales[37])
k12 = hoja.cell(row=12, column=11, value=medidoresiniciales[38])
k16 = hoja.cell(row=16, column=11, value=medidoresiniciales[39])
k20 = hoja.cell(row=20, column=11, value=medidoresiniciales[40])
k24 = hoja.cell(row=24, column=11, value=medidoresiniciales[41])
k28 = hoja.cell(row=28, column=11, value=medidoresiniciales[42])
k32 = hoja.cell(row=32, column=11, value=medidoresiniciales[43])

#Medidores finales
b3 = hoja.cell(row=3, column=2, value=medidoresfinales[0])
b7 = hoja.cell(row=7, column=2, value=medidoresfinales[1])
b11 = hoja.cell(row=11, column=2, value=medidoresfinales[2])
b15 = hoja.cell(row=15, column=2, value=medidoresfinales[3])
b19 = hoja.cell(row=19, column=2, value=medidoresfinales[4])
b23 = hoja.cell(row=23, column=2, value=medidoresfinales[5])
b27 = hoja.cell(row=27, column=2, value=medidoresfinales[6])
b31 = hoja.cell(row=31, column=2, value=medidoresfinales[7])
b35 = hoja.cell(row=35, column=2, value=medidoresfinales[8])
b39 = hoja.cell(row=39, column=2, value=medidoresfinales[9])
b43 = hoja.cell(row=43, column=2, value=medidoresfinales[10])
b47 = hoja.cell(row=47, column=2, value=medidoresfinales[11])
#b43 = hoja.cell(row=43, column=2, value=gncydieseltfinales[44])
#b47 = hoja.cell(row=47, column=2, value=gncydieseltfinales[45])

e3 = hoja.cell(row=3, column=5, value=medidoresfinales[12])
e7 = hoja.cell(row=7, column=5, value=medidoresfinales[13])
e11 = hoja.cell(row=11, column=5, value=medidoresfinales[14])
e15 = hoja.cell(row=15, column=5, value=medidoresfinales[15])
e19 = hoja.cell(row=19, column=5, value=medidoresfinales[16])
e23 = hoja.cell(row=23, column=5, value=medidoresfinales[17])
e27 = hoja.cell(row=27, column=5, value=medidoresfinales[18])
e31 = hoja.cell(row=31, column=5, value=medidoresfinales[19])
e35 = hoja.cell(row=35, column=5, value=medidoresfinales[20])
e39 = hoja.cell(row=39, column=5, value=medidoresfinales[21])
e43 = hoja.cell(row=43, column=5, value=medidoresfinales[22])
e47 = hoja.cell(row=47, column=5, value=medidoresfinales[23])

h3 = hoja.cell(row=3, column=8, value=gncydieseltfinales[24])
h7 = hoja.cell(row=7, column=8, value=gncydieseltfinales[25])
h11 = hoja.cell(row=11, column=8, value=gncydieseltfinales[26])
h15 = hoja.cell(row=15, column=8, value=gncydieseltfinales[27])

h23 = hoja.cell(row=23, column=8, value=medidoresfinales[28])
h27 = hoja.cell(row=27, column=8, value=medidoresfinales[29])
h31 = hoja.cell(row=31, column=8, value=medidoresfinales[30])
h35 = hoja.cell(row=35, column=8, value=medidoresfinales[31])
h39 = hoja.cell(row=39, column=8, value=medidoresfinales[32])
h43 = hoja.cell(row=43, column=8, value=medidoresfinales[33])
h47 = hoja.cell(row=47, column=8, value=medidoresfinales[34])
h51 = hoja.cell(row=51, column=8, value=medidoresfinales[35])

k3 = hoja.cell(row=3, column=11, value=medidoresfinales[36])
k7 = hoja.cell(row=7, column=11, value=medidoresfinales[37])
k11 = hoja.cell(row=11, column=11, value=medidoresfinales[38])
k15 = hoja.cell(row=15, column=11, value=medidoresfinales[39])
k19 = hoja.cell(row=19, column=11, value=medidoresfinales[40])
k23 = hoja.cell(row=23, column=11, value=medidoresfinales[41])
k27 = hoja.cell(row=27, column=11, value=medidoresfinales[42])
k31 = hoja.cell(row=31, column=11, value=medidoresfinales[43])

#Fecha
j47 = hoja.cell(row=47, column=10, value=dia+'/'+mes+'/'+año)

#Ventas de combustibles en Cuenta Corriente
for r28 in hoja:
    try:
        r28 = hoja.cell(row=28, column=18, value=dieselx10)
    except: continue
#r29 = hoja.cell(row=29, column=18, value=dieselgrado2)
for p30 in hoja:
    try:
        p30 = hoja.cell(row=30, column=16, value=quantiumd)
    except: continue
p31 = hoja.cell(row=31, column=16, value=gnc)
p32 = hoja.cell(row=32, column=16, value=quantiump)
p33 = hoja.cell(row=33, column=16, value=energy5000)
p37 = hoja.cell(row=37, column=16, value=axionmix)

#AxionCard
b51 = hoja.cell(row=51, column=2, value=oddieselx10)
e51 = hoja.cell(row=51, column=5, value=odquantiumd)
h55 = hoja.cell(row=55, column=8, value=odquantiump)
k35 = hoja.cell(row=35, column=11, value=odenergy5000)

import os
os.chdir('Z:\excel\A. CAJA DIARIA\\'+año+'\\'+mes)

wb.save('CAJA DIARIA '+dia+'-'+mes+'.xlsx')
